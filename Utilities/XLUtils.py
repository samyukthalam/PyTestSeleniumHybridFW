import openpyxl


def getRowCount(file,sheetName):
    workbook=openpyxl.Workbook(file)
    sheet=workbook[sheetName]
    return sheet.max_row

def getColumnCount(file,sheetName):
    workbook = openpyxl.Workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column

def readExcelData(file,sheetName,rowNum,colNum):
    workbook = openpyxl.Workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum,column=colNum).value

def writeExcelData(file,sheetName,rowNum,colNum,data):
    workbook = openpyxl.Workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value=data
    workbook.save()